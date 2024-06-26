from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.http import HttpResponse
from django.template.loader import get_template
from datetime import datetime
from django.db.models.functions import TruncMonth, ExtractWeekDay

from django.contrib.auth import authenticate, login
from django.db.models import Avg
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from django.urls import reverse
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.tokens import default_token_generator

from django.utils.encoding import force_bytes, force_str
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from rest_framework.authtoken.models import Token
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.hashers import check_password
from django.core.exceptions import ValidationError
from django.core.mail import EmailMultiAlternatives
from django.db.models import Count, Avg, Sum, F
from core.models import Usuario, Estudiante, Admin, Profesor, Clase, Materia, Sesion, Evaluacion, ClaseMateria

from transbank.webpay.webpay_plus.transaction import Transaction
from transbank.common.options import WebpayOptions
from transbank.common.integration_type import IntegrationType
from transbank.common.integration_commerce_codes import IntegrationCommerceCodes
from transbank.common.integration_api_keys import IntegrationApiKeys
# Create your views here.


from django.shortcuts import render


def PaginaPrincipal(request):
    profesores = Profesor.objects.select_related('usuario').prefetch_related('clases_profesor').all()
    materias = Materia.objects.all()

    contexto = {
        "profesores": profesores,
        "user": request.user,
        "materias": materias
    }
    return render(request, 'core/html/PaginaPrincipal.html', contexto)
def ClasesLenguaje(request):
    # Obtener la instancia de la materia 'Lenguaje'
    materia_lenguaje = Materia.objects.get(id_materia=1)

    # Filtrar clases que tienen la materia 'Lenguaje'
    clases = Clase.objects.filter(materias=materia_lenguaje)

    return render(request, 'core/html/ClasesLenguaje.html', {'clases': clases})


def ClasesMatematica(request):
    # Obtener la instancia de la materia 'Matemática'
    materia_matematica = Materia.objects.get(id_materia=2)

    # Filtrar clases que tienen la materia 'Matemática'
    clases = Clase.objects.filter(materias=materia_matematica)

    return render(request, 'core/html/ClasesMatematica.html', {'clases': clases})


def ClasesHistoria(request):
    # Obtener la instancia de la materia 'Historia'
    materia_historia = Materia.objects.get(id_materia=3)

    # Filtrar clases que tienen la materia 'Historia'
    clases = Clase.objects.filter(materias=materia_historia)

    return render(request, 'core/html/ClasesHistoria.html', {'clases': clases})

def Clases(request):
    # Obtener la instancia de la materia 'Historia'


    # Filtrar clases que tienen la materia 'Historia'
    clases = Clase.objects.all()

    return render(request, 'core/html/Clases.html', {'clases': clases})




def Login(request):
    return render(request, 'core/Logueo/Login.html')


def Logueo(request):
    if request.method == 'POST':
        correo = request.POST.get('email1')
        contra = request.POST.get('contra1')

        # Buscar al usuario en el modelo personalizado
        try:
            usuario1 = Usuario.objects.get(email=correo)
        except Usuario.DoesNotExist:
            messages.error(request, 'No se encontró el usuario')
            return redirect('Login')

        # Validar la contraseña ingresada con la contraseña almacenada
        if contra == usuario1.contra:
            # Verificar el tipo de usuario
            if usuario1.tipo_de_usuario == "Profesor":
                try:
                    profesor = Profesor.objects.get(usuario=usuario1)
                    if profesor.estado_de_aprobacion == 'Pendiente':
                        messages.error(request, 'Tu cuenta está pendiente de aprobación. Por favor, espera la aprobación para iniciar sesión.')
                        return redirect('Login')
                except Profesor.DoesNotExist:
                    messages.error(request, 'Tu cuenta de profesor no está configurada correctamente.')
                    return redirect('Login')

            elif usuario1.tipo_de_usuario == "Estudiante":
                try:
                    estudiante = Estudiante.objects.get(usuario=usuario1)
                    if estudiante.estado_solicitud == 'Pendiente':
                        messages.error(request, 'Tu cuenta está pendiente de aprobación por parte de tus padres. Por favor, espera la aprobación para iniciar sesión.')
                        return redirect('Login')
                except Estudiante.DoesNotExist:
                    messages.error(request, 'Tu cuenta de estudiante no está configurada correctamente.')
                    return redirect('Login')

            # Autenticar al usuario utilizando authenticate
            user = authenticate(request, username=usuario1.email, password=usuario1.contra)
            if user is not None:
                login(request, user)
                token, created = Token.objects.get_or_create(user=user)
                # Redirigir según el tipo de usuario
                if usuario1.tipo_de_usuario == "Admin":
                    return redirect('PanelAdmin')
                elif usuario1.tipo_de_usuario == "Estudiante" or usuario1.tipo_de_usuario == "Profesor":
                    return redirect('Perfil')
            else:
                messages.error(request, 'La contraseña es incorrecta')
                return redirect('Login')
        else:
            messages.error(request, 'La contraseña es incorrecta')
            return redirect('Login')

    return redirect('Login')

def Deslogueo(request):
    logout(request)
    return redirect('PaginaPrincipal')

def Agendar (request, id_profesor):
    profe = Profesor.objects.select_related('usuario').get(id_profesor=id_profesor)
    usuario = Usuario.objects.get(email = request.user.username)
    clase = Clase.objects.get(profesor = profe)
    contexto = {
        "profe": profe,
        "clase": clase,
        "usuario": usuario
    }
    return render(request, 'core/html/Agendar.html', contexto)
    
def CambiarContra (request):
    return render(request, 'core/html/CambiarContra.html')

def Solicitudes(request):
    solicitudes = Profesor.objects.filter(estado_de_aprobacion="Pendiente")  # Solo las solicitudes pendientes
    return render(request, 'core/html/Solicitudes.html', {'solicitudes': solicitudes})

# Vista para aceptar una solicitud
def AceptarSolicitud(request, id_solicitud):
    try:
        profesor = Profesor.objects.get(id_profesor=id_solicitud)
        profesor.estado_de_aprobacion = "Aprobado"
        profesor.save()
        send_email(profesor.usuario.email, request, 'aprobado')
        messages.success(request, "Solicitud aceptada con éxito.")
    except Profesor.DoesNotExist:
        messages.error(request, "Solicitud no encontrada.")
    return redirect('Solicitudes')

# Vista para rechazar una solicitud
def RechazarSolicitud(request, id_solicitud):
    try:
        profesor = Profesor.objects.get(id_profesor=id_solicitud)
        usuario = profesor.usuario

        # Eliminar el usuario de Django asociado
        try:
            user = User.objects.get(username=usuario.email)
            user.delete()
        except User.DoesNotExist:
            pass  # Si no existe el usuario en la tabla de User, no hacemos nada

        # Eliminar el profesor y el usuario de la tabla personalizada
        profesor.delete()
        usuario.delete()

        send_email(usuario.email, request, 'rechazado')
        messages.success(request, "Solicitud rechazada y usuario eliminado con éxito.")
    except Profesor.DoesNotExist:
        messages.error(request, "Solicitud no encontrada.")
    except Usuario.DoesNotExist:
        messages.error(request, "Usuario no encontrado.")
    return redirect('Solicitudes')

def DetalleSolicitud(request, id_solicitud):
    profesor = get_object_or_404(Profesor, id_profesor=id_solicitud)
    return render(request, 'core/html/DetalleSolicitud.html', {'profesor': profesor})

def PanelAdmin(request):
    total_usuarios = Usuario.objects.count()
    total_profesores = Profesor.objects.count()
    total_clases = Clase.objects.count()
    total_solicitudes = Profesor.objects.filter(estado_de_aprobacion='pendiente').count()
    ingresos_totales = Sesion.objects.filter(estado_pago=True).aggregate(total=Sum('clase__tarifa_clase'))
    
    sesiones_por_estado = Sesion.objects.values('estado_clase').annotate(count=Count('id_sesion'))
    
    # Consulta para MySQL: Sesiones por Mes
    sesiones_por_mes = Sesion.objects.extra(
        select={'month': "DATE_FORMAT(fechaclase, '%%Y-%%m')"}
    ).values('month').annotate(count=Count('id_sesion')).order_by('month')
    
    # Consulta para MySQL: Sesiones por Día de la Semana
    sesiones_por_dia_semana = Sesion.objects.extra(
        select={'day_of_week': "WEEKDAY(fechaclase)"}
    ).values('day_of_week').annotate(count=Count('id_sesion')).order_by('day_of_week')

    estudiantes_actividades = Estudiante.objects.annotate(count_sesiones=Count('sesion')).order_by('-count_sesiones')[:10]
    profesores_actividades = Profesor.objects.annotate(count_sesiones=Count('sesiones_profesor')).order_by('-count_sesiones')[:10]

    evaluacion_promedio = Evaluacion.objects.aggregate(Avg('valoracion'))['valoracion__avg']

    preferencias_edad = Sesion.objects.values(
        'estudiante__usuario__edad'
    ).annotate(
        avg_prof_edad=Avg('profesor__usuario__edad'),
        pref_prof_genero=Count('profesor__usuario__sexo'),
        count_clases=Count('id_sesion'),
        pref_prof_edad=Avg('profesor__usuario__edad'),
        pref_materia=Count('clase__materias'),
        pref_idioma=Count('clase__idioma_clase')
    ).order_by('estudiante__usuario__edad')
    
    context = {
        'total_usuarios': total_usuarios,
        'total_profesores': total_profesores,
        'total_clases': total_clases,
        'total_solicitudes': total_solicitudes,
        'ingresos_totales': ingresos_totales,
        'sesiones_por_estado': sesiones_por_estado,
        'sesiones_por_mes': sesiones_por_mes,
        'sesiones_por_dia_semana': sesiones_por_dia_semana,
        'estudiantes_actividades': estudiantes_actividades,
        'profesores_actividades': profesores_actividades,
        'evaluacion_promedio': evaluacion_promedio,
        'preferencias_edad': preferencias_edad
    }
    return render(request, 'core/html/PanelAdmin.html', context)

# Vista para exportar datos a Excel

def exportar_excel(request):
    # Obtener todos los datos necesarios para exportar a Excel
    total_usuarios = Usuario.objects.count()
    total_profesores = Profesor.objects.count()
    total_clases = Clase.objects.count()
    total_solicitudes = Profesor.objects.filter(estado_de_aprobacion='pendiente').count()
    ingresos_totales = Sesion.objects.filter(estado_pago=True).aggregate(total=Sum('clase__tarifa_clase'))
    
    sesiones_por_estado = Sesion.objects.values('estado_clase').annotate(count=Count('id_sesion'))
    
    # Consulta para MySQL: Sesiones por Mes
    sesiones_por_mes = Sesion.objects.extra(
        select={'month': "DATE_FORMAT(fechaclase, '%%Y-%%m')"}
    ).values('month').annotate(count=Count('id_sesion')).order_by('month')
    
    # Consulta para MySQL: Sesiones por Día de la Semana
    sesiones_por_dia_semana = Sesion.objects.extra(
        select={'day_of_week': "WEEKDAY(fechaclase)"}
    ).values('day_of_week').annotate(count=Count('id_sesion')).order_by('day_of_week')

    estudiantes_actividades = Estudiante.objects.annotate(count_sesiones=Count('sesion')).order_by('-count_sesiones')[:10]
    profesores_actividades = Profesor.objects.annotate(count_sesiones=Count('sesiones_profesor')).order_by('-count_sesiones')[:10]

    evaluacion_promedio = Evaluacion.objects.aggregate(Avg('valoracion'))['valoracion__avg']

    preferencias_edad = Sesion.objects.values(
        'estudiante__usuario__edad'
    ).annotate(
        avg_prof_edad=Avg('profesor__usuario__edad'),
        pref_prof_genero=Count('profesor__usuario__sexo'),
        count_clases=Count('id_sesion'),
        pref_prof_edad=Avg('profesor__usuario__edad'),
        pref_materia=Count('clase__materias'),
        pref_idioma=Count('clase__idioma_clase')
    ).order_by('estudiante__usuario__edad')
    
    # Crear un DataFrame con los datos
    df_sesiones_por_estado = pd.DataFrame(list(sesiones_por_estado))
    
    # Crear un archivo Excel y escribir los datos en él
    wb = Workbook()
    ws = wb.active
    ws.title = 'Datos de Sesiones'

    # Escribir los datos del DataFrame en la hoja de trabajo
    for r_idx, row in enumerate(df_sesiones_por_estado.iterrows(), start=1):
        ws.cell(row=r_idx + 1, column=1, value=row[1]['estado_clase'])
        ws.cell(row=r_idx + 1, column=2, value=row[1]['count'])

    # Hoja de resumen
    ws_resumen = wb.create_sheet(title='Resumen')
    resumen_data = [
        ['Total de Usuarios', total_usuarios],
        ['Total de Profesores', total_profesores],
        ['Total de Solicitudes Pendientes', total_solicitudes],
        ['Total de Clases', total_clases],
        ['Evaluación Promedio de Clases', evaluacion_promedio],
        ['Ingresos Totales', ingresos_totales['total']]
    ]
    for row in resumen_data:
        ws_resumen.append(row)

    # Hoja de sesiones por mes
    ws_sesiones_por_mes = wb.create_sheet(title='Sesiones por Mes')
    for r_idx, sesion_mes in enumerate(sesiones_por_mes, start=1):
        ws_sesiones_por_mes.cell(row=r_idx, column=1, value=sesion_mes['month'].strftime('%B %Y'))
        ws_sesiones_por_mes.cell(row=r_idx, column=2, value=sesion_mes['count'])

    # Hoja de sesiones por día de la semana
    dias_semana = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
    ws_sesiones_por_dia_semana = wb.create_sheet(title='Sesiones por Día de la Semana')
    for r_idx, sesion_dia in enumerate(sesiones_por_dia_semana, start=1):
        dia_nombre = dias_semana[sesion_dia['day_of_week']]
        ws_sesiones_por_dia_semana.cell(row=r_idx, column=1, value=dia_nombre)
        ws_sesiones_por_dia_semana.cell(row=r_idx, column=2, value=sesion_dia['count'])

    # Guardar el archivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="reporte_admin.xlsx"'
    wb.save(response)

    return response   
    
def PerfilProfe (request):
    return render(request, 'core/html/PerfilProfe.html')

def VistaProfe(request, id_profesor, id_clase):
    # Obtener el profesor y la clase correspondientes
    profe = Profesor.objects.select_related('usuario').get(id_profesor=id_profesor)
    clase = Clase.objects.select_related('profesor').get(id_clase=id_clase)
    cantResenas = Evaluacion.objects.filter(clase_id=id_clase).count()
    avgResena = Evaluacion.objects.filter(clase_id=id_clase).aggregate(promedio=Avg('valoracion'))['promedio']
    
    # Obtener todas las evaluaciones de la clase
    evaluaciones = Evaluacion.objects.filter(clase=clase)

    # Obtener el usuario actual si está autenticado
    usuario_actual = request.user
    estudiante = None
    evaluacion_existente = None

    if usuario_actual.is_authenticated:
        # Si el usuario está autenticado, intenta obtener el estudiante correspondiente y la evaluación existente
        try:
            estudiante = get_object_or_404(Estudiante, usuario__email=usuario_actual.email)
            evaluacion_existente = Evaluacion.objects.filter(
                profesor_id=id_profesor,
                estudiante=estudiante,
                clase_id=id_clase
            ).first()
        except Estudiante.DoesNotExist:
            # Si el usuario autenticado no es un estudiante, establecer evaluacion_existente como None
            pass
    
    # Pasar los datos al contexto
    contexto = {
        "profe": profe,
        "clase": clase,
        "evaluaciones": evaluaciones,
        "evaluacion_existente": evaluacion_existente,
        "cantResenas": cantResenas,
        "Usuario": usuario_actual,
        "avgResena":avgResena 
    }
    
    # Renderizar la plantilla con el contexto
    return render(request, 'core/html/VistaProfe.html', contexto)

def RegistroProfe(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        nombre = request.POST.get('nombre')
        apellido = request.POST.get('apellido')
        edad = request.POST.get('edad')
        sexo = request.POST.get('sexo')
        telefono = request.POST.get('telefono')
        descripcion = request.POST.get('descripcion')
        run = request.POST.get('run')
        foto = request.FILES.get('foto_profe')
        carnet = request.FILES.get('carnet')
        certificado = request.FILES.get('certificado')
        antecedentes = request.FILES.get('antecedentes')
        contra = request.POST.get('contra')

        # Verificar si el correo electrónico ya está en uso
        if Usuario.objects.filter(email=email).exists():
            messages.warning(request, 'El correo ya está en uso')
            return redirect('RegistroProfe')

        # Crear el usuario personalizado
        usuario = Usuario.objects.create(
            email=email,
            nombre=nombre,
            edad=edad,
            sexo = sexo,
            apellido=apellido,
            telefono=telefono,
            contra=contra,
            foto=foto,
            tipo_de_usuario="Profesor"
        )

        # Crear el usuario de Django asociado
        user = User.objects.create_user(
            username=email,
            email=email,
            password=contra,
            first_name=nombre,
            last_name=apellido
        )

        # Crear el profesor asociado al usuario personalizado
        profesor = Profesor.objects.create(
            usuario=usuario,
            antecedentes=antecedentes,
            run=run,
            carnet=carnet,
            certificado = certificado,
            descripcion=descripcion,
            estado_de_aprobacion="Pendiente"
        )

        send_email(email, request, 'registro')  # Enviar el correo de notificación

        messages.success(request, "Registro completado con éxito. Tu cuenta está pendiente de aprobación.")
        return redirect('Login')  # Redirigir después de registrar con éxito

    return render(request, 'core/html/RegistroProfe.html', context={"materias": Materia.objects.all()})

def RegistroEstudiante(request):
    return render(request, 'core/html/RegistroEstudiante.html')

def FormularioEstudiante(request):
    if request.method == 'POST':
        vFoto = request.FILES.get('fotoAlumno')
        vNombre = request.POST.get('nombre')
        vSexo = request.POST.get('sexo')
        vApellido = request.POST.get('apellido')
        vTelefono = request.POST.get('telefono')
        vCorreo = request.POST.get('email')
        vEdad = request.POST.get('edad')
        vCorreoPadre = request.POST.get('correo_padres')
        vClave = request.POST.get('contrasena')
        vNvlEducativo = request.POST.get('NvlEducativo')

        # Verificar si el correo electrónico ya está en uso
        if Usuario.objects.filter(email=vCorreo).exists():
            messages.warning(request, 'El correo ya está en uso')
            return redirect('RegistroEstudiante')

        # Crear el usuario personalizado
        usuario = Usuario.objects.create(
            email=vCorreo,
            nombre=vNombre,
            sexo=vSexo,
            apellido=vApellido,
            telefono=vTelefono,
            contra=vClave,
            edad=vEdad,
            foto=vFoto,
            tipo_de_usuario="Estudiante"
        )

        # Crear el usuario de Django asociado
        user = User.objects.create_user(
            username=vCorreo,
            email=vCorreo,
            password=vClave,
            first_name=vNombre,
            last_name=vApellido
        )

        if int(vEdad) > 18:
            Vestado_solicitud = "Aprobado"
        else:
            Vestado_solicitud = "Pendiente"

        # Crear el estudiante asociado al usuario personalizado
        estudiante = Estudiante.objects.create(
            usuario=usuario,
            correo_padre=vCorreoPadre,
            estado_solicitud=Vestado_solicitud,
            nivel_educativo=vNvlEducativo
        )

        # Enviar correo de validación si el estudiante es menor de edad
        if int(vEdad) < 18 and vCorreoPadre:
            messages.success(request, "Registro completado con éxito. Un correo de validación ha sido enviado a sus padres.")
            send_email(vCorreoPadre, request, 'validacion_estudiante', student_name=estudiante.usuario.nombre, student_id=estudiante.id_estudiante)

        return redirect('Login')

def RegistroAdmin(request):
    return render(request, 'core/html/RegistroAdmin.html')

def FormularioAdmin(request):
    if request.method == 'POST':
        nombre = request.POST.get('nombre')
        apellido = request.POST.get('apellido')
        email = request.POST.get('email')
        contra = request.POST.get('contra')
        telefono = request.POST.get('telefono', '')
        foto = request.FILES.get('foto', None)  # Si hay una imagen, se procesa


        if Usuario.objects.filter(email=email).exists():
            messages.error(request, "Ya existe un usuario con este correo.")
            return render(request, 'core/html/RegistroAdmin.html')

        try:
            # Crear usuario personalizado
            usuario = Usuario(
                email=email,
                nombre=nombre,
                apellido=apellido,
                contra=contra,
                tipo_de_usuario='Admin',
                telefono=telefono,
                foto=foto,
            )
            usuario.save()

            # Crear usuario Django para autenticación
            user = User.objects.create_user(
                username=email,
                email=email,
                password=contra,
                first_name=nombre,
                last_name=apellido
            )

            # Crear el perfil de Admin
            admin = Admin(
                usuario=usuario,
                nombre=nombre,
            )
            admin.save()

            messages.success(request, "Administrador creado con éxito.")
            return redirect('PanelAdmin')  # Redirigir al panel del administrador

        except Exception as e:
            messages.error(request, f"Error al crear el administrador: {e}")

def CambiarContra(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        try:
            user = User.objects.get(email=email)
            send_email(email, request, 'cambiar')  # Pasar el tipo 'cambiar'
            messages.success(request, 'Se ha enviado un correo para cambiar tu contraseña.')
            return redirect('CambiarContra')  # Redirigir a la misma página para mostrar el mensaje
        except User.DoesNotExist:
            messages.error(request, 'El correo electrónico no está registrado.')
    
    return render(request, 'core/html/CambiarContra.html')

def solicitar_cambio_contra(request, tipo):
    if request.method == 'POST':
        email = request.POST.get('email')
        try:
            user = User.objects.get(email=email)
            send_email(email, request, 'restablecer')
            messages.success(request, f'Se ha enviado un correo para {tipo} tu contraseña.')
            return redirect('Login')
        except User.DoesNotExist:
            messages.error(request, 'El correo electrónico no está registrado.')
    
    return render(request, 'core/html/solicitar_cambio_contra.html', {'tipo': tipo})

def send_email(email, request, tipo, student_name=None, student_id=None):
    context = {
        'email': email,
        'request': request,
        'tipo': tipo,
    }

    if tipo == 'validacion_estudiante' and student_name and student_id:
        context.update({
            'student_name': student_name,
            'student_id': student_id
        })

    subject = {
        'cambiar': 'Cambio de Contraseña',
        'restablecer': 'Restablecimiento de Contraseña',
        'validacion_estudiante': 'Validación de Registro de Estudiante',
        'aprobado': 'Registro Aprobado',
        'rechazado': 'Registro Rechazado',
        'notificacion': 'Correo de notificación'
    }.get(tipo, 'Correo de notificación')

    template = get_template('core/html/Correo_generico.html')
    content = template.render(context)

    mail = EmailMultiAlternatives(
        subject,
        'Correo de notificación',
        settings.EMAIL_HOST_USER,
        [email]
    )

    mail.attach_alternative(content, 'text/html')
    mail.send()

def reset_password(request, email):
    if request.method == 'POST': 
        new_password = request.POST.get('password')
        
        try:
            user = User.objects.get(email=email)
            usuario = get_object_or_404(Usuario, email=email)
            
            # Actualizar la contraseña en ambas tablas
            user.set_password(new_password)
            user.save()
            
            usuario.contra = new_password
            usuario.save()

            messages.success(request, 'Tu contraseña ha sido cambiada exitosamente.')
            return redirect('Login')
        except User.DoesNotExist:
            messages.error(request, 'El correo electrónico no está registrado.')
    
    return render(request, 'core/html/reset_password.html', {'email': email})
    
@login_required
def Perfil(request):
    usuario = Usuario.objects.get(email=request.user.username)
    try:
        profe = Profesor.objects.get(usuario=usuario)
    except Profesor.DoesNotExist:
        profe = None
    
    return render(request, 'core/html/Perfil.html', {'usuario': usuario, 'profe': profe})

def ModificarPerfil(request):
    usuario = request.user
    usuarioActivo = Usuario.objects.get(email=usuario.email)
    profe = Profesor.objects.get(usuario=usuarioActivo)

    if request.method == 'POST':
        vFoto = request.FILES.get('fotoPerfil')
        vNombre = request.POST.get('nombre')
        vApellido = request.POST.get('apellido')
        vTelefono = request.POST.get('telefono')
        vEdad = request.POST.get('edad')
        vDescripcion = request.POST.get('descripcion')

        try:
            # Actualizar campos del Usuario existente
            usuarioActivo.nombre = vNombre
            usuarioActivo.apellido = vApellido
            usuarioActivo.telefono = vTelefono
            usuarioActivo.edad = vEdad
            if vFoto:
                usuarioActivo.foto = vFoto
            usuarioActivo.save()

            # Actualizar campos del Profesor
            profe.descripcion = vDescripcion
            profe.save()

            print('Perfil actualizado exitosamente')
        except :
             print('Error al actualizar perfil')

        return redirect('Perfil')  # Ajusta 'Perfil' al nombre de tu vista de perfil


@login_required
def ListaUsuarios(request):
    usuarios = Usuario.objects.all()
    return render(request, 'core/html/ListaUsuarios.html', {'usuarios': usuarios})

@login_required
def ListaClases(request):
    usuario = request.user
    usuario1 = Usuario.objects.get(email=usuario)
    profe = Profesor.objects.get(usuario=usuario1)
    clases = Clase.objects.filter(profesor=profe).distinct()
    return render(request, 'core/html/ListaClases.html', {'clases': clases})

def EliminarUsuario(request, usuario_id):
    try:
        # Buscar el usuario por ID
        usuario = Usuario.objects.get(id_usuario=usuario_id)

        # Eliminar el usuario
        usuario.delete()

        # Mensaje de éxito
        messages.success(request, f"Usuario {usuario.nombre} {usuario.apellido} eliminado con éxito.")

    except Usuario.DoesNotExist:
        # Manejar caso donde el usuario no exista
        messages.error(request, "Usuario no encontrado.")

    except Exception as e:
        # Manejar otros errores inesperados
        messages.error(request, f"Error al eliminar usuario: {e}")

    # Redirigir a la lista de usuarios
    return redirect('ListaUsuarios')

@login_required
def EliminarClase(request, clase_id):
    usuario = request.user
    usuario1 = Usuario.objects.get(email=usuario)
    profe = Profesor.objects.get(usuario=usuario1)
    
    try:
        # Obtén la clase que el profesor quiere eliminar
        clase = Clase.objects.get(id_clase=clase_id, profesor=profe)
        clase.delete()
        messages.success(request, "Clase eliminada con éxito.")
    except Clase.DoesNotExist:
        messages.error(request, "Clase no encontrada o no tienes permiso para eliminar esta clase.")
    except Exception as e:
        messages.error(request, f"Error al eliminar clase: {e}")
    
    return redirect('ClasesProfe')

@login_required
def VerClase(request, clase_id):
    clase = get_object_or_404(Sesion, id_sesion=clase_id)
    return render(request, 'core/html/VerClase.html', {'clase': clase})

@login_required
def FormClase(request):
    if request.method == 'POST':
        titulo = request.POST.get('titulo')
        descripcion = request.POST.get('descripcion')
        precio = request.POST.get('precio')
        materia = request.POST.get('materia')
        idioma = request.POST.get('idioma')
        
        usuario = request.user
        
        try:
            usuario1 = Usuario.objects.get(email=usuario)
            idusuario = Profesor.objects.get(usuario=usuario1)
            materiaid = Materia.objects.get(id_materia=materia)
            
            clase = Clase(
                nombre_clase=titulo,
                tarifa_clase=precio,
                descripcion_clase=descripcion,
                profesor=idusuario,
                idioma = idioma
            )
            clase.save()
            
            claseMateria = ClaseMateria(
                clase_id=clase.id_clase,
                materia_id=materiaid.id_materia
            )
            claseMateria.save()
            
            return redirect('ClasesProfe')
        except Exception as e:
            messages.error(request, f"Error al crear la clase: {e}")
            return redirect('CrearClase')

@login_required
def CrearClase(request):
    materias = Materia.objects.all()
    return render(request, 'core/html/FormClase.html', {'materias': materias })

@login_required
def ClasesProfe(request):
    usuario = request.user
    usuario1 = Usuario.objects.get(email=usuario)
    profe = Profesor.objects.get(usuario=usuario1)
    clases = Clase.objects.filter(profesor=profe).distinct()
    return render(request, 'core/html/VerClases.html', {'clases' : clases})

@login_required
def EditarClase(request, id_clase):
    if request.method == 'POST':
        usuario = request.user
        usuario1 = Usuario.objects.get(email=usuario)
        profe = Profesor.objects.get(usuario=usuario1)
        
        clase = get_object_or_404(Clase, id_clase=id_clase, profesor=profe)
        
        clase.nombre_clase = request.POST.get('nombreClase')
        clase.tarifa_clase = request.POST.get('tarifaClase')
        clase.descripcion_clase = request.POST.get('descripcionClase')
        
        clase.save()
        return redirect('ClasesProfe')
    else:
        return render(request, 'PaginaPrincipal')          

def Agendar(request, id_profesor, id_clase):
    profe = Profesor.objects.select_related('usuario').get(id_profesor=id_profesor)
    usuario = Usuario.objects.get(email=request.user.username)
    estudiante = Estudiante.objects.get(usuario=usuario)
    clase = Clase.objects.select_related('profesor').get(id_clase=id_clase)
    cantResenas = Evaluacion.objects.filter(clase_id=id_clase).count()
    avgResena = Evaluacion.objects.filter(clase_id=id_clase).aggregate(promedio=Avg('valoracion'))['promedio']
    try:
        sesion = Sesion.objects.get(estudiante = estudiante)
        sesionAgendada = Sesion.objects.filter(estudiante=estudiante, clase=clase).first()
    except:
        sesion = None
        sesionAgendada = None

    # Verificar si existe una sesión agendada

    
    contexto = {
        "profe": profe,
        "clase": clase,
        "usuario": usuario,
        "estudiante": estudiante,
        "sesion": sesion,
        "sesionAgendada": sesionAgendada,
        "avgResena": avgResena,
        "cantResenas": cantResenas
    }
    
    return render(request, 'core/html/Agendar.html', contexto)

tx = Transaction(WebpayOptions(IntegrationCommerceCodes.WEBPAY_PLUS, IntegrationApiKeys.WEBPAY, IntegrationType.TEST))

def pagar(request, id_sesion):
    print(f"Llegó una solicitud de pago para la sesión: {id_sesion}")
    sesion = get_object_or_404(Sesion, id_sesion=id_sesion)

    # Datos de la transacción
    buy_order = str(sesion.id_sesion)
    session_id = str(sesion.estudiante.id_estudiante)
    amount = sesion.clase.tarifa_clase
    return_url = request.build_absolute_uri(reverse('retorno'))

    response = tx.create(buy_order, session_id, amount, return_url)
    print(f"Respuesta de tx.create: {response}")

    token = response.get('token')
    url = response.get('url')

    if not token or not url:
        return HttpResponse("Error en la creación de la transacción")

    return render(request, 'core/html/pagos/pago_formulario.html', {'url': url, 'token': token})


def retorno(request):
    token_ws = request.GET.get('token_ws', None)
    if token_ws:
        response = tx.commit(token_ws)
        if response['response_code'] == 0 and response['status'] == 'AUTHORIZED':
            # Transacción aprobada
            buy_order = response.get('buy_order')  # Obtén el identificador de la sesión
            try:
                sesion = Sesion.objects.get(id_sesion=buy_order)
                sesion.estado_pago = 1  # Actualiza el estado de pago
                sesion.save()  # Guarda los cambios
            except Sesion.DoesNotExist:
                return HttpResponse("Error: Sesión no encontrada")

            return render(request, 'core/html/pagos/pago_exitoso.html', {'response': response})
        else:
            # Transacción rechazada
            return render(request, 'core/html/pagos/pago_rechazado.html', {'response': response})
    else:
        return HttpResponse("Error: no se recibió token_ws")

def FormularioAgendar(request):
     if request.method == 'POST':
        mensaje = request.POST.get('mensaje')
        fecha_str = request.POST.get('datepicker')
        hora_str = request.POST.get('timepicker')
        telefono = request.POST.get('telefono')
        id_profesor = request.POST.get('id_profesor')
        usuario_actual = request.user
        usuario = get_object_or_404(Usuario, email=usuario_actual.email)
        estudiante = get_object_or_404(Estudiante, usuario=usuario)
        id_clase = request.POST.get('id_clase') 
        id_estudiante = estudiante.id_estudiante

        # Combinar fecha y hora
        datetime_str = f"{fecha_str} {hora_str}"
        fecha_hora = datetime.strptime(datetime_str, '%d/%m/%Y %H:%M')

        nueva_sesion = Sesion.objects.create(
            mensaje=mensaje,
            fechaclase=fecha_hora,
            contacto=telefono,
            profesor_id=id_profesor,
            estudiante_id=id_estudiante,
            clase_id=id_clase
        )
        print("ID del estudiante:", id_estudiante)

        return render (request, 'core/html/Agendar.html')

def Calificar(request, id_profesor, id_clase):
    # Obtener el usuario actual y el estudiante asociado
    usuario_actual = request.user
    estudiante = get_object_or_404(Estudiante, usuario__email=usuario_actual.email)

    # Intentar obtener la evaluación existente
    try:
        evaluacion_existente = Evaluacion.objects.get(
            profesor_id=id_profesor,
            estudiante=estudiante,
            clase_id=id_clase
        )
    except: evaluacion_existente = None

    if request.method == 'POST':
        # Obtener la calificación y el comentario del formulario
        calificacion = request.POST.get('calificacion')
        comentario = request.POST.get('comentario')

        # Si ya existe una evaluación, actualizarla; de lo contrario, crear una nueva
        if evaluacion_existente:
            evaluacion_existente.valoracion = calificacion
            evaluacion_existente.recomendacion = comentario
            evaluacion_existente.save()
        else:
            Evaluacion.objects.create(
                recomendacion=comentario,
                valoracion=calificacion,
                profesor_id=id_profesor,
                estudiante=estudiante,
                clase_id=id_clase
            )

        # Redireccionar a la vista del profesor después de la calificación
        return redirect('VistaProfe', id_profesor=id_profesor, id_clase=id_clase)

def ValidacionPapasView(request, student_id):
    estudiante = get_object_or_404(Estudiante, id_estudiante=student_id)
    return render(request, 'core/html/ValidacionPapas.html', {'estudiante': estudiante})

def ValidacionPapas(request, student_id, decision):
    try:
        estudiante = Estudiante.objects.get(id_estudiante=student_id)
        usuario = estudiante.usuario  # Obtener el usuario relacionado antes de eliminar el estudiante

        if decision == 'aceptar':
            estudiante.estado_solicitud = "Aprobado"
            estudiante.save()
            messages.success(request, "Solicitud aceptada con éxito.")
        else:
            usuario_email = usuario.email  # Guardar el email del usuario antes de eliminarlo
            estudiante.delete()  # Eliminar primero el estudiante

            # Eliminar el usuario de Django asociado
            try:
                user = User.objects.get(username=usuario.email)
                user.delete()
            except User.DoesNotExist:
                pass  # Si no existe el usuario en la tabla de User, no hacemos nada

            usuario.delete()  # Eliminar el usuario personalizado

            # Enviar correo de rechazo al estudiante
            send_email(usuario_email, request, 'rechazado')
            messages.success(request, "Solicitud rechazada y usuario eliminado con éxito.")

    except Estudiante.DoesNotExist:
        messages.error(request, "Solicitud no encontrada.")
    except Usuario.DoesNotExist:
        messages.error(request, "Usuario no encontrado.")

    return redirect('Login')

def CorreoPapas(request):
    return render(request, 'core/html/CorreoPapas.html')

def ValidacionCorreoPapa(request):
    correo = request.POST.get('correo')
    try:
        estudiante = Estudiante.objects.get(correo_padre=correo)
        if estudiante:
            return redirect('ValidacionPapas', correo=correo)
    except Estudiante.DoesNotExist:
        messages.error(request, 'Correo no encontrado')
        return redirect('CorreoPapas')
    
def AceptarSolicitudEstudiante(request, id_estudiante):
    try:
        estudiante = Estudiante.objects.get(id_estudiante=id_estudiante)
        estudiante.estado_solicitud = "Aprobado"
        estudiante.save()

        # Enviar correo de aprobación al estudiante
        send_email(estudiante.usuario.email, request, 'aprobado')

        # Enviar correo de confirmación al padre
        send_email(estudiante.correo_padre, request, 'notificacion', student_name=estudiante.usuario.nombre)

        messages.success(request, "Solicitud aceptada con éxito.")
    except Estudiante.DoesNotExist:
        messages.error(request, "Solicitud no encontrada.")
    return redirect('Login')

def RechazarSolicitudEstudiante(request, id_estudiante):
    try:
        estudiante = Estudiante.objects.get(id_estudiante=id_estudiante)
        usuario = estudiante.usuario

        # Eliminar el usuario de Django asociado
        try:
            user = User.objects.get(username=usuario.email)
            user.delete()
        except User.DoesNotExist:
            pass  # Si no existe el usuario en la tabla de User, no hacemos nada

        # Eliminar el estudiante y el usuario personalizado
        estudiante.delete()
        usuario.delete()

        # Enviar correo de rechazo al estudiante
        send_email(usuario.email, request, 'rechazado')

        messages.success(request, "Solicitud rechazada y usuario eliminado con éxito.")
    except Estudiante.DoesNotExist:
        messages.error(request, "Solicitud no encontrada.")
    except Usuario.DoesNotExist:
        messages.error(request, "Usuario no encontrado.")
    return redirect('Solicitudes')


def SolicitudClase(request):
    # Obtener todas las solicitudes de sesión pendientes
    solicitudes = Sesion.objects.filter(estado_clase="Pendiente")

    nombres_clases = []

    for sesion in solicitudes:
        nombre_clase = sesion.clase.nombre_clase
        nombres_clases.append(nombre_clase)

    return render(request, 'core/html/SolicitudClase.html', {'solicitudes': solicitudes, 'nombres_clases': nombres_clases})

def AceptarSolicitudClase(request,id_sesion):
    try:
        sesion = get_object_or_404(Sesion, id_sesion=id_sesion)
        sesion.estado_clase = "Aprobado"
        sesion.save()
    except Sesion.DoesNotExist:
        messages.error(request, "Solicitud no encontrada.")
    messages.success(request, "Solicitud aceptada con éxito.")
    return redirect('SolicitudClase')

def RechazarSolicitudClase(request,id_sesion):
    try:
        sesion = get_object_or_404(Sesion, id_sesion=id_sesion)
        sesion.delete()
    except Sesion.DoesNotExist:
        pass  # Si no existe el usuario en la tabla de User, no hacemos nada
    return redirect('SolicitudClase')


def DetalleSolicitudClase(request, id_sesion):
    sesion = get_object_or_404(Sesion, id_sesion=id_sesion)
    usuario = sesion.estudiante.usuario
    clase = sesion.clase
    return render(request, 'core/html/DetalleSolicitudClase.html', {'sesion': sesion,'usuario': usuario, 'clase': clase})


def VerSesiones(request):
    usuario=request.user
    UsuarioActual = Usuario.objects.get(email = usuario )
    profesor = Profesor.objects.get( usuario = UsuarioActual)
    sesiones = Sesion.objects.filter(profesor = profesor, estado_realizacion = "Pendiente")

    contexto = {
        'sesiones': sesiones
    }

    return render(request, 'core/html/VerSesiones.html', contexto )

def ClaseRealizada(request,sesion_id):
    sesion = get_object_or_404(Sesion, id_sesion=sesion_id)
    sesion.estado_realizacion = "Realizada"
    sesion.save()
    return redirect('VerSesiones')  # Asume que tienes una URL llamada 'ver_sesiones' que apunta a tu vista VerSesiones   








    # def SolicitudesClase(request):
    #     solicitudes = Profesor.objects.filter(estado_clase="Pendiente")  # Solo las solicitudes pendientes
    # return render(request, 'core/html/VerClases.html', {'solicitudes': solicitudes})

    # def AceptarClase(request, id_solicitud):
    # try:
    #     profesor = Profesor.objects.get(id_profesor=id_solicitud)
    #     profesor.estado_de_aprobacion = "Aprobado"  # Cambiar el estado a Aprobado
    #     profesor.save()  # Guardar los cambios
    #     messages.success(request, "Solicitud aceptada con éxito.")
    # except Profesor.DoesNotExist:
    #     messages.error(request, "Solicitud no encontrada.")
    # return redirect('Solicitudes')  # Redirigir a la página de solicitudes
